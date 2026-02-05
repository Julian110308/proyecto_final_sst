from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


class ExcelReporteGenerator:
    """
    Generador de reportes Excel para el sistema SST
    """

    # Estilos comunes
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    HEADER_FILL = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    SUBHEADER_FILL = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
    ALT_ROW_FILL = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center')

    @staticmethod
    def _aplicar_estilo_header(ws, row, num_cols, fill=None):
        """Aplica estilo de encabezado a una fila"""
        fill = fill or ExcelReporteGenerator.HEADER_FILL
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = ExcelReporteGenerator.HEADER_FONT
            cell.fill = fill
            cell.alignment = ExcelReporteGenerator.CENTER_ALIGN
            cell.border = ExcelReporteGenerator.BORDER

    @staticmethod
    def _aplicar_estilo_datos(ws, start_row, end_row, num_cols):
        """Aplica estilo a filas de datos"""
        for row in range(start_row, end_row + 1):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = ExcelReporteGenerator.BORDER
                cell.alignment = ExcelReporteGenerator.CENTER_ALIGN
                if (row - start_row) % 2 == 1:
                    cell.fill = ExcelReporteGenerator.ALT_ROW_FILL

    @staticmethod
    def _ajustar_anchos(ws, col_widths):
        """Ajusta el ancho de las columnas"""
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    @staticmethod
    def generar_reporte_aforo(datos):
        """
        Genera Excel para reporte de aforo
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Aforo"

        # Título
        ws.merge_cells('A1:D1')
        ws['A1'] = 'REPORTE DE AFORO - CENTRO MINERO SENA'
        ws['A1'].font = Font(bold=True, size=14, color='2C3E50')
        ws['A1'].alignment = ExcelReporteGenerator.CENTER_ALIGN

        ws.merge_cells('A2:D2')
        ws['A2'] = f"Periodo: {datos['periodo_inicio']} a {datos['periodo_fin']}"
        ws['A2'].alignment = ExcelReporteGenerator.CENTER_ALIGN

        # Resumen General
        ws['A4'] = 'RESUMEN GENERAL'
        ws['A4'].font = Font(bold=True, size=12)

        resumen = [
            ['Indicador', 'Valor'],
            ['Total de Ingresos', datos['total_ingresos']],
            ['Aforo Máximo Permitido', datos['aforo_maximo_permitido']],
            ['Porcentaje de Uso', f"{datos['porcentaje_uso']}%"],
            ['Hora Pico', f"{datos['hora_pico']}:00 hrs" if datos['hora_pico'] else 'N/A'],
            ['Total en Hora Pico', datos['total_hora_pico']],
            ['Tiempo Promedio Permanencia',
             f"{datos['tiempo_promedio_permanencia_minutos']} min" if datos['tiempo_promedio_permanencia_minutos'] else 'N/A'],
        ]

        start_row = 5
        for i, row_data in enumerate(resumen):
            for j, value in enumerate(row_data):
                ws.cell(row=start_row + i, column=j + 1, value=value)

        ExcelReporteGenerator._aplicar_estilo_header(ws, 5, 2, ExcelReporteGenerator.SUBHEADER_FILL)
        ExcelReporteGenerator._aplicar_estilo_datos(ws, 6, 11, 2)

        # Aforo por Rol
        if datos['aforo_por_rol']:
            ws['A14'] = 'AFORO POR ROL'
            ws['A14'].font = Font(bold=True, size=12)

            ws.cell(row=15, column=1, value='Rol')
            ws.cell(row=15, column=2, value='Total de Ingresos')
            ExcelReporteGenerator._aplicar_estilo_header(ws, 15, 2,
                PatternFill(start_color='2ECC71', end_color='2ECC71', fill_type='solid'))

            for i, item in enumerate(datos['aforo_por_rol']):
                ws.cell(row=16 + i, column=1, value=item['rol'])
                ws.cell(row=16 + i, column=2, value=item['total'])

            ExcelReporteGenerator._aplicar_estilo_datos(ws, 16, 15 + len(datos['aforo_por_rol']), 2)

        # Aforo Diario (nueva hoja)
        if datos['aforo_diario']:
            ws_diario = wb.create_sheet(title="Aforo Diario")
            ws_diario['A1'] = 'AFORO DIARIO'
            ws_diario['A1'].font = Font(bold=True, size=12)

            ws_diario.cell(row=2, column=1, value='Fecha')
            ws_diario.cell(row=2, column=2, value='Total Ingresos')
            ExcelReporteGenerator._aplicar_estilo_header(ws_diario, 2, 2)

            for i, item in enumerate(datos['aforo_diario']):
                fecha = item['fecha']
                if hasattr(fecha, 'strftime'):
                    fecha = fecha.strftime('%Y-%m-%d')
                ws_diario.cell(row=3 + i, column=1, value=str(fecha))
                ws_diario.cell(row=3 + i, column=2, value=item['total'])

            ExcelReporteGenerator._aplicar_estilo_datos(ws_diario, 3, 2 + len(datos['aforo_diario']), 2)
            ExcelReporteGenerator._ajustar_anchos(ws_diario, [15, 15])

        # Pie de página
        last_row = ws.max_row + 2
        ws.cell(row=last_row, column=1,
                value=f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        ws.cell(row=last_row, column=1).font = Font(italic=True, size=9, color='7F8C8D')

        ExcelReporteGenerator._ajustar_anchos(ws, [30, 25])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generar_reporte_incidentes(datos):
        """
        Genera Excel para reporte de incidentes/emergencias
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Incidentes"

        # Título
        ws.merge_cells('A1:D1')
        ws['A1'] = 'REPORTE DE INCIDENTES Y EMERGENCIAS - CENTRO MINERO SENA'
        ws['A1'].font = Font(bold=True, size=14, color='E74C3C')
        ws['A1'].alignment = ExcelReporteGenerator.CENTER_ALIGN

        ws.merge_cells('A2:D2')
        ws['A2'] = f"Periodo: {datos['periodo_inicio']} a {datos['periodo_fin']}"
        ws['A2'].alignment = ExcelReporteGenerator.CENTER_ALIGN

        # Resumen General
        ws['A4'] = 'RESUMEN GENERAL'
        ws['A4'].font = Font(bold=True, size=12)

        resumen = [
            ['Indicador', 'Valor'],
            ['Total de Emergencias', datos['total_emergencias']],
            ['Personas Afectadas', datos['total_personas_afectadas']],
            ['Evacuaciones Requeridas', datos['evacuaciones_requeridas']],
            ['Porcentaje Resueltas', f"{datos['porcentaje_resueltas']}%"],
            ['Tiempo Promedio Respuesta',
             f"{datos['tiempo_promedio_respuesta_minutos']} min" if datos['tiempo_promedio_respuesta_minutos'] else 'N/A'],
        ]

        start_row = 5
        for i, row_data in enumerate(resumen):
            for j, value in enumerate(row_data):
                ws.cell(row=start_row + i, column=j + 1, value=value)

        ExcelReporteGenerator._aplicar_estilo_header(ws, 5, 2,
            PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid'))
        ExcelReporteGenerator._aplicar_estilo_datos(ws, 6, 10, 2)

        # Por Tipo
        if datos['por_tipo']:
            ws['A13'] = 'INCIDENTES POR TIPO'
            ws['A13'].font = Font(bold=True, size=12)

            ws.cell(row=14, column=1, value='Tipo')
            ws.cell(row=14, column=2, value='Total')
            ExcelReporteGenerator._aplicar_estilo_header(ws, 14, 2,
                PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid'))

            for i, item in enumerate(datos['por_tipo']):
                ws.cell(row=15 + i, column=1, value=item.get('tipo_nombre', item.get('tipo', 'N/A')))
                ws.cell(row=15 + i, column=2, value=item['total'])

            ExcelReporteGenerator._aplicar_estilo_datos(ws, 15, 14 + len(datos['por_tipo']), 2)

        # Por Estado
        if datos['por_estado']:
            current_row = ws.max_row + 3
            ws.cell(row=current_row, column=1, value='INCIDENTES POR ESTADO')
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)

            ws.cell(row=current_row + 1, column=1, value='Estado')
            ws.cell(row=current_row + 1, column=2, value='Total')
            ExcelReporteGenerator._aplicar_estilo_header(ws, current_row + 1, 2,
                PatternFill(start_color='9B59B6', end_color='9B59B6', fill_type='solid'))

            for i, item in enumerate(datos['por_estado']):
                ws.cell(row=current_row + 2 + i, column=1, value=item['estado'])
                ws.cell(row=current_row + 2 + i, column=2, value=item['total'])

            ExcelReporteGenerator._aplicar_estilo_datos(ws, current_row + 2,
                current_row + 1 + len(datos['por_estado']), 2)

        # Pie de página
        last_row = ws.max_row + 2
        ws.cell(row=last_row, column=1,
                value=f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        ws.cell(row=last_row, column=1).font = Font(italic=True, size=9, color='7F8C8D')

        ExcelReporteGenerator._ajustar_anchos(ws, [35, 20])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generar_reporte_asistencia(datos):
        """
        Genera Excel para reporte de asistencia
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Asistencia"

        # Título
        ws.merge_cells('A1:E1')
        ws['A1'] = 'REPORTE DE ASISTENCIA - CENTRO MINERO SENA'
        ws['A1'].font = Font(bold=True, size=14, color='16A085')
        ws['A1'].alignment = ExcelReporteGenerator.CENTER_ALIGN

        ws.merge_cells('A2:E2')
        ws['A2'] = f"Ficha: {datos['ficha']} | Periodo: {datos['periodo_inicio']} a {datos['periodo_fin']}"
        ws['A2'].alignment = ExcelReporteGenerator.CENTER_ALIGN

        # Resumen General
        ws['A4'] = 'RESUMEN GENERAL'
        ws['A4'].font = Font(bold=True, size=12)

        resumen = [
            ['Indicador', 'Valor'],
            ['Total de Aprendices', datos['total_aprendices']],
            ['Días Totales del Periodo', datos['dias_totales']],
            ['Promedio de Asistencia', f"{datos['promedio_asistencia']}%"],
        ]

        start_row = 5
        for i, row_data in enumerate(resumen):
            for j, value in enumerate(row_data):
                ws.cell(row=start_row + i, column=j + 1, value=value)

        ExcelReporteGenerator._aplicar_estilo_header(ws, 5, 2,
            PatternFill(start_color='16A085', end_color='16A085', fill_type='solid'))
        ExcelReporteGenerator._aplicar_estilo_datos(ws, 6, 8, 2)

        # Detalle por Aprendiz
        ws['A11'] = 'DETALLE POR APRENDIZ'
        ws['A11'].font = Font(bold=True, size=12)

        headers = ['Nombre', 'Documento', 'Días Asistió', 'Días Totales', '% Asistencia', 'Estado']
        for j, header in enumerate(headers):
            ws.cell(row=12, column=j + 1, value=header)
        ExcelReporteGenerator._aplicar_estilo_header(ws, 12, len(headers),
            PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid'))

        for i, aprendiz in enumerate(datos['aprendices']):
            ws.cell(row=13 + i, column=1, value=aprendiz['nombre'])
            ws.cell(row=13 + i, column=2, value=aprendiz['documento'])
            ws.cell(row=13 + i, column=3, value=aprendiz['dias_asistio'])
            ws.cell(row=13 + i, column=4, value=aprendiz['dias_totales'])
            ws.cell(row=13 + i, column=5, value=f"{aprendiz['porcentaje_asistencia']}%")
            ws.cell(row=13 + i, column=6, value=aprendiz['estado'])

            # Color por estado
            estado_cell = ws.cell(row=13 + i, column=6)
            if aprendiz['estado'] == 'APROBADO':
                estado_cell.fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
            else:
                estado_cell.fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')

        ExcelReporteGenerator._aplicar_estilo_datos(ws, 13, 12 + len(datos['aprendices']), len(headers))

        # Pie de página
        last_row = ws.max_row + 2
        ws.cell(row=last_row, column=1,
                value=f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        ws.cell(row=last_row, column=1).font = Font(italic=True, size=9, color='7F8C8D')

        ExcelReporteGenerator._ajustar_anchos(ws, [30, 15, 12, 12, 12, 12])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generar_reporte_seguridad(datos):
        """
        Genera Excel para reporte de seguridad
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Seguridad"

        # Título
        ws.merge_cells('A1:D1')
        ws['A1'] = 'REPORTE DE SEGURIDAD - CENTRO MINERO SENA'
        ws['A1'].font = Font(bold=True, size=14, color='34495E')
        ws['A1'].alignment = ExcelReporteGenerator.CENTER_ALIGN

        ws.merge_cells('A2:D2')
        ws['A2'] = f"Periodo: {datos['periodo_inicio']} a {datos['periodo_fin']}"
        ws['A2'].alignment = ExcelReporteGenerator.CENTER_ALIGN

        # Estado de Equipamiento
        ws['A4'] = 'ESTADO DEL EQUIPAMIENTO DE SEGURIDAD'
        ws['A4'].font = Font(bold=True, size=12)

        resumen = [
            ['Indicador', 'Valor'],
            ['Total de Equipos', datos['equipamiento_total']],
            ['Equipos Operativos', datos['equipamiento_operativo']],
            ['En Mantenimiento', datos['equipamiento_mantenimiento']],
            ['Fuera de Servicio', datos['equipamiento_fuera_servicio']],
            ['Porcentaje Operativo', f"{datos['porcentaje_operativo']}%"],
        ]

        start_row = 5
        for i, row_data in enumerate(resumen):
            for j, value in enumerate(row_data):
                ws.cell(row=start_row + i, column=j + 1, value=value)

        ExcelReporteGenerator._aplicar_estilo_header(ws, 5, 2,
            PatternFill(start_color='34495E', end_color='34495E', fill_type='solid'))
        ExcelReporteGenerator._aplicar_estilo_datos(ws, 6, 10, 2)

        # Equipamiento por Tipo
        if datos['equipamiento_por_tipo']:
            ws['A13'] = 'EQUIPAMIENTO POR TIPO'
            ws['A13'].font = Font(bold=True, size=12)

            headers = ['Tipo', 'Total', 'Operativos']
            for j, header in enumerate(headers):
                ws.cell(row=14, column=j + 1, value=header)
            ExcelReporteGenerator._aplicar_estilo_header(ws, 14, 3,
                PatternFill(start_color='2980B9', end_color='2980B9', fill_type='solid'))

            for i, item in enumerate(datos['equipamiento_por_tipo']):
                ws.cell(row=15 + i, column=1, value=item['tipo'])
                ws.cell(row=15 + i, column=2, value=item['total'])
                ws.cell(row=15 + i, column=3, value=item['operativo'])

            ExcelReporteGenerator._aplicar_estilo_datos(ws, 15, 14 + len(datos['equipamiento_por_tipo']), 3)

        # Emergencias por Tipo
        if datos['emergencias_por_tipo']:
            current_row = ws.max_row + 3
            ws.cell(row=current_row, column=1, value='EMERGENCIAS POR TIPO')
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)

            ws.cell(row=current_row + 1, column=1, value='Tipo')
            ws.cell(row=current_row + 1, column=2, value='Total')
            ExcelReporteGenerator._aplicar_estilo_header(ws, current_row + 1, 2,
                PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid'))

            for i, item in enumerate(datos['emergencias_por_tipo']):
                ws.cell(row=current_row + 2 + i, column=1, value=item.get('tipo_nombre', 'N/A'))
                ws.cell(row=current_row + 2 + i, column=2, value=item['total'])

            ExcelReporteGenerator._aplicar_estilo_datos(ws, current_row + 2,
                current_row + 1 + len(datos['emergencias_por_tipo']), 2)

        # Tendencias Mensuales (nueva hoja)
        if datos.get('tendencias_mensuales'):
            ws_tend = wb.create_sheet(title="Tendencias")
            ws_tend['A1'] = 'TENDENCIAS MENSUALES DE EMERGENCIAS'
            ws_tend['A1'].font = Font(bold=True, size=12)

            ws_tend.cell(row=2, column=1, value='Mes')
            ws_tend.cell(row=2, column=2, value='Total Emergencias')
            ExcelReporteGenerator._aplicar_estilo_header(ws_tend, 2, 2)

            for i, item in enumerate(datos['tendencias_mensuales']):
                ws_tend.cell(row=3 + i, column=1, value=item['mes'])
                ws_tend.cell(row=3 + i, column=2, value=item['total'])

            ExcelReporteGenerator._aplicar_estilo_datos(ws_tend, 3, 2 + len(datos['tendencias_mensuales']), 2)
            ExcelReporteGenerator._ajustar_anchos(ws_tend, [15, 20])

        # Pie de página
        last_row = ws.max_row + 2
        ws.cell(row=last_row, column=1,
                value=f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        ws.cell(row=last_row, column=1).font = Font(italic=True, size=9, color='7F8C8D')

        ExcelReporteGenerator._ajustar_anchos(ws, [30, 15, 15])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generar_reporte_emergencias(datos):
        """
        Genera Excel para reporte detallado de emergencias
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Emergencias"

        # Título
        ws.merge_cells('A1:G1')
        ws['A1'] = 'REPORTE DETALLADO DE EMERGENCIAS - CENTRO MINERO SENA'
        ws['A1'].font = Font(bold=True, size=14, color='C0392B')
        ws['A1'].alignment = ExcelReporteGenerator.CENTER_ALIGN

        ws.merge_cells('A2:G2')
        ws['A2'] = f"Periodo: {datos['periodo_inicio']} a {datos['periodo_fin']}"
        ws['A2'].alignment = ExcelReporteGenerator.CENTER_ALIGN

        # Resumen
        ws['A4'] = 'RESUMEN'
        ws['A4'].font = Font(bold=True, size=12)

        resumen = [
            ['Indicador', 'Valor'],
            ['Total Emergencias', datos.get('total_emergencias', 0)],
            ['Resueltas', datos.get('resueltas', 0)],
            ['En Atención', datos.get('en_atencion', 0)],
            ['Tiempo Promedio Respuesta', f"{datos.get('tiempo_promedio_respuesta', 0)} min"],
            ['Tiempo Promedio Resolución', f"{datos.get('tiempo_promedio_resolucion', 0)} min"],
        ]

        for i, row_data in enumerate(resumen):
            for j, value in enumerate(row_data):
                ws.cell(row=5 + i, column=j + 1, value=value)

        ExcelReporteGenerator._aplicar_estilo_header(ws, 5, 2,
            PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid'))
        ExcelReporteGenerator._aplicar_estilo_datos(ws, 6, 10, 2)

        # Listado de Emergencias
        if datos.get('emergencias'):
            ws['A13'] = 'LISTADO DE EMERGENCIAS'
            ws['A13'].font = Font(bold=True, size=12)

            headers = ['Fecha', 'Tipo', 'Estado', 'Ubicación', 'Reportado Por',
                      'Tiempo Respuesta', 'Afectados']
            for j, header in enumerate(headers):
                ws.cell(row=14, column=j + 1, value=header)
            ExcelReporteGenerator._aplicar_estilo_header(ws, 14, len(headers))

            for i, emerg in enumerate(datos['emergencias']):
                ws.cell(row=15 + i, column=1, value=emerg.get('fecha', ''))
                ws.cell(row=15 + i, column=2, value=emerg.get('tipo', ''))
                ws.cell(row=15 + i, column=3, value=emerg.get('estado', ''))
                ws.cell(row=15 + i, column=4, value=emerg.get('ubicacion', ''))
                ws.cell(row=15 + i, column=5, value=emerg.get('reportado_por', ''))
                ws.cell(row=15 + i, column=6, value=f"{emerg.get('tiempo_respuesta', 0)} min")
                ws.cell(row=15 + i, column=7, value=emerg.get('personas_afectadas', 0))

            ExcelReporteGenerator._aplicar_estilo_datos(ws, 15, 14 + len(datos['emergencias']), len(headers))

        ExcelReporteGenerator._ajustar_anchos(ws, [18, 20, 15, 25, 20, 18, 12])

        # Pie de página
        last_row = ws.max_row + 2
        ws.cell(row=last_row, column=1,
                value=f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        ws.cell(row=last_row, column=1).font = Font(italic=True, size=9, color='7F8C8D')

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
