# Vistas SIMPLES para el módulo de incidentes
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from .models import Incidente
from .forms import IncidenteForm
# Servicio centralizado de notificaciones
from usuarios.services import NotificacionService


def _calcular_sla(incidente):
    """Devuelve el estado SLA de un incidente no resuelto."""
    if incidente.estado in ['RESUELTO', 'CERRADO']:
        return 'ok'
    horas = (timezone.now() - incidente.fecha_reporte).total_seconds() / 3600
    if horas > 72:
        return 'critico'
    if horas > 24:
        return 'vencido'
    if horas > 8:
        return 'proximo'
    return 'ok'


# Vista SIMPLE: Listar todos los incidentes
@login_required
def listar_incidentes(request):
    """
    Muestra todos los incidentes con filtros avanzados y SLA.
    Administradores y brigada ven todos; los demás solo los suyos.
    """
    if request.user.rol in ['ADMINISTRATIVO', 'BRIGADA']:
        incidentes = Incidente.objects.select_related('reportado_por', 'asignado_a').all()
    else:
        incidentes = Incidente.objects.select_related('reportado_por', 'asignado_a').filter(
            reportado_por=request.user
        )

    # --- Filtros ---
    q = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '')
    area = request.GET.get('area', '')
    gravedad = request.GET.get('gravedad', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if q:
        from django.db.models import Q
        incidentes = incidentes.filter(
            Q(titulo__icontains=q) | Q(descripcion__icontains=q) | Q(persona_afectada__icontains=q)
        )
    if estado:
        incidentes = incidentes.filter(estado=estado)
    if area:
        incidentes = incidentes.filter(area_incidente=area)
    if gravedad:
        incidentes = incidentes.filter(gravedad=gravedad)
    if fecha_desde:
        try:
            from datetime import datetime
            incidentes = incidentes.filter(
                fecha_reporte__date__gte=datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            )
        except ValueError:
            pass
    if fecha_hasta:
        try:
            from datetime import datetime
            incidentes = incidentes.filter(
                fecha_reporte__date__lte=datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    # Contar por estado (sobre el queryset filtrado)
    total = incidentes.count()
    reportados = incidentes.filter(estado='REPORTADO').count()
    en_proceso = incidentes.filter(estado__in=['EN_REVISION', 'EN_PROCESO']).count()
    resueltos = incidentes.filter(estado__in=['RESUELTO', 'CERRADO']).count()
    criticos_sla = incidentes.exclude(estado__in=['RESUELTO', 'CERRADO']).filter(
        fecha_reporte__lte=timezone.now() - timedelta(hours=72)
    ).count()

    # Anotar SLA en cada incidente para la tabla
    incidentes_list = list(incidentes)
    for inc in incidentes_list:
        inc.sla_estado = _calcular_sla(inc)

    # Paginación: 15 incidentes por página
    paginator = Paginator(incidentes_list, 15)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Choices para filtros en template
    filtros_activos = any([q, estado, area, gravedad, fecha_desde, fecha_hasta])

    context = {
        'incidentes': page_obj,
        'page_obj': page_obj,
        'total': total,
        'reportados': reportados,
        'en_proceso': en_proceso,
        'resueltos': resueltos,
        'criticos_sla': criticos_sla,
        # Filtros aplicados
        'q': q,
        'estado_filtro': estado,
        'area_filtro': area,
        'gravedad_filtro': gravedad,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'filtros_activos': filtros_activos,
        # Choices para los selects
        'estado_choices': Incidente.ESTADO_CHOICES,
        'area_choices': Incidente.AREA_CHOICES,
        'gravedad_choices': Incidente.GRAVEDAD_CHOICES,
    }

    return render(request, 'reportes/incidentes_lista.html', context)


# Vista SIMPLE: Crear nuevo incidente
@login_required
def crear_incidente(request):
    """
    Formulario para reportar un nuevo incidente
    Cualquier usuario autenticado puede reportar
    """
    if request.method == 'POST':
        form = IncidenteForm(request.POST, request.FILES)

        if form.is_valid():
            # Guardar pero no commit todavía
            incidente = form.save(commit=False)

            # Asignar el usuario que reporta
            incidente.reportado_por = request.user

            # Si es crítico, asignar automáticamente a admin
            if incidente.gravedad == 'CRITICA':
                from usuarios.models import Usuario
                admin = Usuario.objects.filter(rol='ADMINISTRATIVO').first()
                if admin:
                    incidente.asignado_a = admin

            # Guardar
            incidente.save()

            # Notificar a administrativos (y a instructores si es critico)
            NotificacionService.notificar_incidente_creado(incidente)

            # Generar alarma para brigada si gravedad es ALTA o CRITICA
            if incidente.gravedad in ['ALTA', 'CRITICA']:
                NotificacionService.notificar_alarma_incidente(incidente)

            messages.success(request, f'Incidente "{incidente.titulo}" reportado exitosamente. Se ha generado una alerta al personal correspondiente.')
            return redirect('listar_incidentes')
    else:
        form = IncidenteForm()

    context = {
        'form': form
    }

    return render(request, 'reportes/incidente_form.html', context)


# Vista SIMPLE: Ver detalle de un incidente
@login_required
def detalle_incidente(request, pk):
    """
    Muestra los detalles completos de un incidente
    """
    incidente = get_object_or_404(Incidente, pk=pk)

    # Verificar permisos: solo el que reportó, admin o brigada pueden ver
    if request.user.rol not in ['ADMINISTRATIVO', 'BRIGADA'] and incidente.reportado_por != request.user:
        messages.error(request, 'No tienes permiso para ver este incidente.')
        return redirect('listar_incidentes')

    # Calcular tiempo transcurrido desde el reporte
    ahora = timezone.now()
    delta = ahora - incidente.fecha_reporte
    horas_transcurridas = delta.total_seconds() / 3600
    sla_estado = _calcular_sla(incidente)

    context = {
        'incidente': incidente,
        'horas_transcurridas': round(horas_transcurridas, 1),
        'sla_estado': sla_estado,
    }

    return render(request, 'reportes/incidente_detalle.html', context)


# Vista SIMPLE: Actualizar estado (admin y brigada)
@login_required
def actualizar_incidente(request, pk):
    """
    Permite al administrador o a la brigada actualizar el estado y acciones del incidente.
    Notifica automáticamente al usuario que reportó cuando el estado cambia.
    """
    if request.user.rol not in ['ADMINISTRATIVO', 'BRIGADA']:
        messages.error(request, 'No tienes permiso para actualizar incidentes.')
        return redirect('listar_incidentes')

    incidente = get_object_or_404(Incidente, pk=pk)

    if request.method == 'POST':
        # Guardar estado anterior para detectar cambios
        estado_anterior = incidente.estado

        # Actualizar estado
        estado = request.POST.get('estado')
        acciones = request.POST.get('acciones_tomadas')
        asignado_a_id = request.POST.get('asignado_a')
        marcar_resuelto = request.POST.get('marcar_resuelto')

        if estado:
            incidente.estado = estado

        if acciones:
            incidente.acciones_tomadas = acciones

        # Asignar a usuario (solo ADMINISTRATIVO puede reasignar)
        if request.user.rol == 'ADMINISTRATIVO':
            if asignado_a_id:
                from usuarios.models import Usuario
                incidente.asignado_a = Usuario.objects.get(id=asignado_a_id)
            elif asignado_a_id == '':
                incidente.asignado_a = None

        # Si se marca como resuelto, guardar fecha
        if marcar_resuelto and not incidente.fecha_resolucion:
            incidente.fecha_resolucion = timezone.now()

        incidente.save()

        # Notificar al reportador si el estado cambió
        if estado and estado_anterior != incidente.estado and incidente.reportado_por:
            from usuarios.models import Notificacion
            estado_labels = dict(Incidente.ESTADO_CHOICES)
            nuevo_estado_label = estado_labels.get(incidente.estado, incidente.estado)
            detalle_acciones = f' Acciones tomadas: {acciones[:100]}' if acciones else ''
            prioridad = 'ALTA' if incidente.estado in ['RESUELTO', 'CERRADO'] else 'MEDIA'
            Notificacion.objects.create(
                destinatario=incidente.reportado_por,
                titulo=f'Actualización de tu incidente: {incidente.titulo}',
                mensaje=f'El estado cambió de "{estado_labels.get(estado_anterior, estado_anterior)}" a "{nuevo_estado_label}".{detalle_acciones}',
                tipo='INCIDENTE',
                prioridad=prioridad,
                url_relacionada=f'/reportes/incidentes/{incidente.pk}/',
            )

        messages.success(request, 'Incidente actualizado exitosamente.')
        return redirect('detalle_incidente', pk=pk)

    # Obtener usuarios disponibles para asignar (solo ADMINISTRATIVO puede ver/cambiar asignación)
    usuarios_disponibles = []
    if request.user.rol == 'ADMINISTRATIVO':
        from usuarios.models import Usuario
        usuarios_disponibles = Usuario.objects.filter(
            activo=True,
            rol__in=['ADMINISTRATIVO', 'INSTRUCTOR', 'BRIGADA']
        )

    context = {
        'incidente': incidente,
        'usuarios_disponibles': usuarios_disponibles,
        'puede_reasignar': request.user.rol == 'ADMINISTRATIVO',
    }

    return render(request, 'reportes/incidente_actualizar.html', context)


# Vista: Exportar incidentes a Excel
@login_required
def exportar_incidentes_excel(request):
    """Exporta la lista de incidentes (con filtros activos) a Excel."""
    if request.user.rol in ['ADMINISTRATIVO', 'BRIGADA']:
        incidentes = Incidente.objects.select_related('reportado_por', 'asignado_a').all()
    else:
        incidentes = Incidente.objects.select_related('reportado_por', 'asignado_a').filter(
            reportado_por=request.user
        )

    # Aplicar los mismos filtros que la vista de lista
    q = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '')
    area = request.GET.get('area', '')
    gravedad = request.GET.get('gravedad', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if q:
        from django.db.models import Q
        incidentes = incidentes.filter(
            Q(titulo__icontains=q) | Q(descripcion__icontains=q)
        )
    if estado:
        incidentes = incidentes.filter(estado=estado)
    if area:
        incidentes = incidentes.filter(area_incidente=area)
    if gravedad:
        incidentes = incidentes.filter(gravedad=gravedad)
    if fecha_desde:
        try:
            from datetime import datetime
            incidentes = incidentes.filter(
                fecha_reporte__date__gte=datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            )
        except ValueError:
            pass
    if fecha_hasta:
        try:
            from datetime import datetime
            incidentes = incidentes.filter(
                fecha_reporte__date__lte=datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            )
        except ValueError:
            pass

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Incidentes'

    # Estilos
    verde = PatternFill('solid', fgColor='39A900')
    blanco_bold = Font(bold=True, color='FFFFFF')
    gris = PatternFill('solid', fgColor='F5F5F5')
    borde_fino = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    # Fila de título
    ws.merge_cells('A1:K1')
    titulo_cell = ws['A1']
    titulo_cell.value = f'Reporte de Incidentes - Centro Minero SENA - {timezone.now().strftime("%d/%m/%Y %H:%M")}'
    titulo_cell.font = Font(bold=True, size=13, color='FFFFFF')
    titulo_cell.fill = verde
    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    # Encabezados
    headers = [
        'ID', 'Título', 'Tipo', 'Área', 'Gravedad', 'Estado',
        'Fecha Incidente', 'Fecha Reporte', 'Fecha Resolución',
        'Reportado Por', 'Asignado A',
    ]
    ws.append(headers)
    for col, cell in enumerate(ws[2], 1):
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2E7D32')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = borde_fino

    ws.row_dimensions[2].height = 16

    # Datos
    colores_gravedad = {'CRITICA': 'FFEBEE', 'ALTA': 'FFF8E1', 'MEDIA': 'E3F2FD', 'BAJA': 'F1F8E9'}
    for i, inc in enumerate(incidentes, 3):
        row = [
            inc.id,
            inc.titulo,
            inc.get_tipo_display(),
            inc.get_area_incidente_display(),
            inc.get_gravedad_display(),
            inc.get_estado_display(),
            inc.fecha_incidente.strftime('%d/%m/%Y %H:%M') if inc.fecha_incidente else '',
            inc.fecha_reporte.strftime('%d/%m/%Y %H:%M') if inc.fecha_reporte else '',
            inc.fecha_resolucion.strftime('%d/%m/%Y %H:%M') if inc.fecha_resolucion else 'Sin resolver',
            inc.reportado_por.get_full_name() or inc.reportado_por.username,
            inc.asignado_a.get_full_name() if inc.asignado_a else 'Sin asignar',
        ]
        ws.append(row)
        color_fila = colores_gravedad.get(inc.gravedad, 'FFFFFF')
        fill = PatternFill('solid', fgColor=color_fila) if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        for cell in ws[i]:
            cell.fill = fill
            cell.border = borde_fino
            cell.alignment = Alignment(vertical='center', wrap_text=False)

    # Ancho de columnas
    anchos = [6, 40, 18, 18, 12, 14, 18, 18, 18, 22, 22]
    for col_idx, ancho in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = ancho

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fecha_str = timezone.now().strftime('%Y%m%d_%H%M')
    response['Content-Disposition'] = f'attachment; filename="incidentes_{fecha_str}.xlsx"'
    wb.save(response)
    return response
