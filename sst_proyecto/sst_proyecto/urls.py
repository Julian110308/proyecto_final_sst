"""
URL configuration for sst_proyecto project.
"""

from django.contrib import admin
from django.urls import path, include
from django.conf import settings
from django.conf.urls.static import static
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import views as auth_views
from django.views.generic import RedirectView
from drf_spectacular.views import SpectacularAPIView, SpectacularSwaggerView, SpectacularRedocView
from usuarios.permissions import rol_requerido, excluir_visitantes
from usuarios.login_view import custom_login_view, visitante_login_view


# Páginas de error personalizadas
def error_404(request, exception=None):
    return render(request, "404.html", status=404)


def error_500(request):
    return render(request, "500.html", status=500)


handler404 = error_404
handler500 = error_500

# Vistas para servir SW y manifest desde la raíz (requerido para PWA)
import os
from django.http import HttpResponse


def _get_static_dir():
    if hasattr(settings, "STATICFILES_DIRS") and settings.STATICFILES_DIRS:
        return settings.STATICFILES_DIRS[0]
    return settings.STATIC_ROOT


def sw_view(request):
    filepath = os.path.join(_get_static_dir(), "sw.js")
    with open(filepath, "r", encoding="utf-8") as f:
        response = HttpResponse(f.read(), content_type="application/javascript")
        response["Service-Worker-Allowed"] = "/"
        return response


def manifest_view(request):
    filepath = os.path.join(_get_static_dir(), "manifest.json")
    with open(filepath, "r", encoding="utf-8") as f:
        return HttpResponse(f.read(), content_type="application/manifest+json")


# Vista principal de dashboard que redirige según el rol
@login_required
def dashboard_view(request):
    """
    Vista principal que redirige al dashboard específico según el rol del usuario
    CON DATOS REALES DE LA BASE DE DATOS
    """
    from usuarios.models import Usuario, Visitante
    from control_acceso.models import RegistroAcceso, ConfiguracionAforo
    from django.db.models import Count
    from django.utils import timezone
    from datetime import timedelta

    usuario = request.user

    # Mapeo de roles a templates
    dashboard_templates = {
        "COORDINADOR_SST": "dashboard/coordinador_sst.html",
        "APRENDIZ": "dashboard/aprendiz.html",
        "INSTRUCTOR": "dashboard/instructor.html",
        "ADMINISTRATIVO": "dashboard/administrativo.html",
        "VIGILANCIA": "dashboard/vigilancia.html",
        "BRIGADA": "dashboard/brigada.html",
        "VISITANTE": "dashboard/visitante.html",
    }

    # Si es COORDINADOR_SST, construir contexto específico y retornar
    if usuario.rol == "COORDINADOR_SST":
        from django.db.models import Count

        pendientes = Usuario.objects.filter(estado_cuenta="PENDIENTE").count()
        total_activos = Usuario.objects.filter(estado_cuenta="ACTIVO", activo=True).count()
        total_bloqueados = Usuario.objects.filter(estado_cuenta="BLOQUEADO").count()
        brigada_activa = Usuario.objects.filter(es_brigada=True, activo=True).count()
        por_rol = list(Usuario.objects.filter(activo=True).values("rol").annotate(total=Count("id")).order_by("rol"))
        return render(
            request,
            "dashboard/coordinador_sst.html",
            {
                "usuario": usuario,
                "pendientes": pendientes,
                "total_activos": total_activos,
                "total_bloqueados": total_bloqueados,
                "brigada_activa": brigada_activa,
                "por_rol": por_rol,
            },
        )

    # Obtener el template según el rol
    template = dashboard_templates.get(usuario.rol, "dashboard.html")

    # CALCULAR DATOS REALES
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)

    # Total de usuarios registrados
    total_usuarios = Usuario.objects.filter(activo=True).count()

    # Usuarios nuevos este mes
    usuarios_mes = Usuario.objects.filter(fecha_registro__gte=inicio_mes, activo=True).count()

    # Personas actualmente en el centro (tienen ingreso pero no egreso)
    personas_en_centro = RegistroAcceso.objects.filter(
        tipo="INGRESO", fecha_hora_egreso__isnull=True, fecha_hora_ingreso__date=hoy
    ).count()

    # Ingresos totales hoy
    ingresos_hoy = RegistroAcceso.objects.filter(tipo="INGRESO", fecha_hora_ingreso__date=hoy).count()

    # Visitantes hoy
    visitantes_hoy = Visitante.objects.filter(fecha_visita=hoy).count()

    # Configuración de aforo
    try:
        config_aforo = ConfiguracionAforo.objects.filter(activo=True).first()
        aforo_maximo = config_aforo.aforo_maximo if config_aforo else 2000
    except Exception:
        aforo_maximo = 2000

    # Porcentaje de ocupación
    porcentaje_ocupacion = round((personas_en_centro / aforo_maximo) * 100, 1) if aforo_maximo > 0 else 0

    # Últimos 7 días de registros (para gráficas) - una sola query agrupada
    from django.db.models.functions import TruncDate

    fecha_inicio_7 = hoy - timedelta(days=6)
    registros_por_dia = dict(
        RegistroAcceso.objects.filter(
            tipo="INGRESO", fecha_hora_ingreso__date__gte=fecha_inicio_7, fecha_hora_ingreso__date__lte=hoy
        )
        .annotate(dia=TruncDate("fecha_hora_ingreso"))
        .values("dia")
        .annotate(cantidad=Count("id"))
        .values_list("dia", "cantidad")
    )
    ultimos_7_dias = []
    for i in range(6, -1, -1):
        fecha = hoy - timedelta(days=i)
        ultimos_7_dias.append({"fecha": fecha.strftime("%d/%m"), "cantidad": registros_por_dia.get(fecha, 0)})

    # Últimos 5 accesos
    ultimos_accesos = (
        RegistroAcceso.objects.select_related("usuario").filter(tipo="INGRESO").order_by("-fecha_hora_ingreso")[:5]
    )

    # Contexto con datos reales
    context = {
        "usuario": usuario,
        "rol": usuario.rol,
        "permisos": usuario.get_permissions(),
        # Métricas principales
        "total_usuarios": total_usuarios,
        "usuarios_mes": usuarios_mes,
        "personas_en_centro": personas_en_centro,
        "ingresos_hoy": ingresos_hoy,
        "visitantes_hoy": visitantes_hoy,
        "aforo_maximo": aforo_maximo,
        "porcentaje_ocupacion": porcentaje_ocupacion,
        # Datos para gráficas
        "ultimos_7_dias": ultimos_7_dias,
        "ultimos_accesos": ultimos_accesos,
    }

    # Datos adicionales para INSTRUCTOR
    if usuario.rol == "INSTRUCTOR":
        ficha_instructor = usuario.ficha or ""
        programa_instructor = usuario.programa_formacion or ""

        filtro_aprendices = {"rol": "APRENDIZ", "activo": True}
        filtro_acceso = {"tipo": "INGRESO", "fecha_hora_ingreso__date": hoy, "usuario__rol": "APRENDIZ"}

        if ficha_instructor:
            filtro_aprendices["ficha"] = ficha_instructor
            filtro_acceso["usuario__ficha"] = ficha_instructor
        elif programa_instructor:
            filtro_aprendices["programa_formacion"] = programa_instructor
            filtro_acceso["usuario__programa_formacion"] = programa_instructor

        total_aprendices = Usuario.objects.filter(**filtro_aprendices).count()
        aprendices_presentes = RegistroAcceso.objects.filter(**filtro_acceso).values("usuario_id").distinct().count()
        ultimos_accesos_aprendices = (
            RegistroAcceso.objects.select_related("usuario").filter(**filtro_acceso).order_by("-fecha_hora_ingreso")[:5]
        )

        if ficha_instructor:
            sublabel_aprendices = f"Ficha {ficha_instructor}"
            sublabel_presentes = f"Ficha {ficha_instructor} en el centro"
        elif programa_instructor:
            sublabel_aprendices = programa_instructor
            sublabel_presentes = "De mi programa en el centro"
        else:
            sublabel_aprendices = "Todos los aprendices SENA"
            sublabel_presentes = "Sin filtro asignado"

        context["total_aprendices"] = total_aprendices
        context["aprendices_presentes"] = aprendices_presentes
        context["ultimos_accesos_aprendices"] = ultimos_accesos_aprendices
        context["programa_instructor"] = programa_instructor
        context["ficha_instructor"] = ficha_instructor
        context["sublabel_aprendices"] = sublabel_aprendices
        context["sublabel_presentes"] = sublabel_presentes

        # Estadísticas por ficha (la del instructor o su programa)
        filtro_ficha_base = {"rol": "APRENDIZ", "activo": True, "ficha__isnull": False}
        filtro_acceso_ficha = {
            "tipo": "INGRESO",
            "fecha_hora_ingreso__date": hoy,
            "usuario__rol": "APRENDIZ",
            "usuario__ficha__isnull": False,
        }
        if ficha_instructor:
            filtro_ficha_base["ficha"] = ficha_instructor
            filtro_acceso_ficha["usuario__ficha"] = ficha_instructor
        elif programa_instructor:
            filtro_ficha_base["programa_formacion"] = programa_instructor
            filtro_acceso_ficha["usuario__programa_formacion"] = programa_instructor

        presentes_por_ficha = dict(
            RegistroAcceso.objects.filter(**filtro_acceso_ficha)
            .exclude(usuario__ficha="")
            .values("usuario__ficha")
            .annotate(presentes=Count("usuario_id", distinct=True))
            .values_list("usuario__ficha", "presentes")
        )
        fichas_stats = list(
            Usuario.objects.filter(**filtro_ficha_base)
            .exclude(ficha="")
            .values("ficha")
            .annotate(total=Count("id"))
            .order_by("-total")[:5]
        )
        for ficha in fichas_stats:
            presentes = presentes_por_ficha.get(ficha["ficha"], 0)
            ficha["presentes"] = presentes
            ficha["porcentaje"] = round((presentes / ficha["total"]) * 100) if ficha["total"] > 0 else 0
        context["fichas_stats"] = fichas_stats

        # Gráfica 7 días filtrada
        filtro_7dias = {
            "tipo": "INGRESO",
            "fecha_hora_ingreso__date__gte": fecha_inicio_7,
            "fecha_hora_ingreso__date__lte": hoy,
            "usuario__rol": "APRENDIZ",
        }
        if ficha_instructor:
            filtro_7dias["usuario__ficha"] = ficha_instructor
        elif programa_instructor:
            filtro_7dias["usuario__programa_formacion"] = programa_instructor
        registros_instructor_por_dia = dict(
            RegistroAcceso.objects.filter(**filtro_7dias)
            .annotate(dia=TruncDate("fecha_hora_ingreso"))
            .values("dia")
            .annotate(cantidad=Count("id"))
            .values_list("dia", "cantidad")
        )
        ultimos_7_dias_instructor = []
        for i in range(6, -1, -1):
            fecha = hoy - timedelta(days=i)
            ultimos_7_dias_instructor.append(
                {"fecha": fecha.strftime("%d/%m"), "cantidad": registros_instructor_por_dia.get(fecha, 0)}
            )
        context["ultimos_7_dias"] = ultimos_7_dias_instructor

        from reportes.models import Incidente as IncidenteInstructor

        context["incidentes_mes"] = IncidenteInstructor.objects.filter(fecha_reporte__date__gte=inicio_mes).count()

        # Estado actual del instructor (si él mismo está en el centro hoy)
        registro_instructor = (
            RegistroAcceso.objects.filter(
                usuario=usuario, tipo="INGRESO", fecha_hora_egreso__isnull=True, fecha_hora_ingreso__date=hoy
            )
            .order_by("-fecha_hora_ingreso")
            .first()
        )
        context["en_centro"] = registro_instructor is not None
        context["registro_activo"] = registro_instructor

    # Datos adicionales para BRIGADA
    if usuario.rol == "BRIGADA":
        from reportes.models import Incidente
        from django.utils import timezone as tz

        incidentes_recientes = Incidente.objects.select_related("reportado_por").order_by("-fecha_reporte")[:10]
        incidentes_pendientes = Incidente.objects.exclude(estado__in=["RESUELTO", "CERRADO"]).count()
        incidentes_total_mes = Incidente.objects.filter(fecha_reporte__date__gte=inicio_mes).count()
        incidentes_criticos_sla = (
            Incidente.objects.exclude(estado__in=["RESUELTO", "CERRADO"])
            .filter(fecha_reporte__lte=tz.now() - timedelta(hours=72))
            .count()
        )
        context["incidentes_recientes"] = incidentes_recientes
        context["incidentes_pendientes"] = incidentes_pendientes
        context["incidentes_total_mes"] = incidentes_total_mes
        context["incidentes_criticos_sla"] = incidentes_criticos_sla

    # Datos adicionales para ADMINISTRATIVO
    if usuario.rol == "ADMINISTRATIVO":
        from reportes.models import Incidente
        from django.utils import timezone as tz
        from django.db.models import Count

        incidentes_pendientes = Incidente.objects.exclude(estado__in=["RESUELTO", "CERRADO"]).count()
        incidentes_total_mes = Incidente.objects.filter(fecha_reporte__date__gte=inicio_mes).count()
        incidentes_criticos = Incidente.objects.filter(gravedad="CRITICA", estado="REPORTADO").count()
        incidentes_criticos_sla = (
            Incidente.objects.exclude(estado__in=["RESUELTO", "CERRADO"])
            .filter(fecha_reporte__lte=tz.now() - timedelta(hours=72))
            .count()
        )
        # Por gravedad para gráfica de dona
        incidentes_por_gravedad = list(
            Incidente.objects.values("gravedad").annotate(total=Count("id")).order_by("gravedad")
        )
        context["incidentes_pendientes"] = incidentes_pendientes
        context["incidentes_total_mes"] = incidentes_total_mes
        context["incidentes_criticos"] = incidentes_criticos
        context["incidentes_criticos_sla"] = incidentes_criticos_sla
        context["incidentes_por_gravedad"] = incidentes_por_gravedad

    # Datos adicionales para APRENDIZ
    if usuario.rol == "APRENDIZ":
        mis_accesos = RegistroAcceso.objects.filter(
            usuario=usuario, tipo="INGRESO", fecha_hora_ingreso__date__gte=inicio_mes
        ).order_by("-fecha_hora_ingreso")[:5]
        mis_ingresos_mes = RegistroAcceso.objects.filter(
            usuario=usuario, tipo="INGRESO", fecha_hora_ingreso__date__gte=inicio_mes
        ).count()
        context["mis_accesos"] = mis_accesos
        context["mis_ingresos_mes"] = mis_ingresos_mes

        # Estado actual: ¿está el aprendiz en el centro ahora mismo?
        registro_activo = (
            RegistroAcceso.objects.filter(
                usuario=usuario, tipo="INGRESO", fecha_hora_egreso__isnull=True, fecha_hora_ingreso__date=hoy
            )
            .order_by("-fecha_hora_ingreso")
            .first()
        )
        context["en_centro"] = registro_activo is not None
        context["registro_activo"] = registro_activo

        # Datos para gráfica de asistencia mensual (últimos 30 días) - una sola query
        fecha_inicio_30 = hoy - timedelta(days=29)
        dias_con_asistencia = set(
            RegistroAcceso.objects.filter(
                usuario=usuario,
                tipo="INGRESO",
                fecha_hora_ingreso__date__gte=fecha_inicio_30,
                fecha_hora_ingreso__date__lte=hoy,
            )
            .annotate(dia=TruncDate("fecha_hora_ingreso"))
            .values_list("dia", flat=True)
            .distinct()
        )
        asistencia_mensual = []
        for i in range(29, -1, -1):
            fecha = hoy - timedelta(days=i)
            asistencia_mensual.append(
                {
                    "fecha": fecha.strftime("%d"),
                    "fecha_completa": fecha.strftime("%d/%m"),
                    "asistio": 1 if fecha in dias_con_asistencia else 0,
                }
            )
        context["asistencia_mensual"] = asistencia_mensual
        context["dias_asistidos_mes"] = len(dias_con_asistencia)

        # Contactos de emergencia
        from emergencias.models import ContactoExterno

        contactos_emergencia = ContactoExterno.objects.filter(activo=True).order_by("orden_contacto")[:5]
        context["contactos_emergencia"] = contactos_emergencia

    return render(request, template, context)


# ==============================================
# VISTA DE PERFIL (Todos los roles)
# ==============================================


@login_required
def mi_perfil_view(request):
    """
    Vista de perfil del usuario autenticado.
    Permite ver y editar datos personales.
    """
    from django.contrib import messages as django_messages

    usuario = request.user

    if request.method == "POST":
        # Campos editables por el usuario
        usuario.first_name = request.POST.get("first_name", usuario.first_name).strip()
        usuario.last_name = request.POST.get("last_name", usuario.last_name).strip()
        usuario.email = request.POST.get("email", usuario.email).strip()
        usuario.telefono = request.POST.get("telefono", usuario.telefono).strip()
        usuario.telefono_emergencia = request.POST.get("telefono_emergencia", usuario.telefono_emergencia).strip()
        usuario.contacto_emergencia = request.POST.get("contacto_emergencia", usuario.contacto_emergencia).strip()

        # Preferencias de notificaciones externas
        usuario.recibir_notif_email = "recibir_notif_email" in request.POST
        usuario.callmebot_apikey = request.POST.get("callmebot_apikey", "").strip()

        # Programa y ficha — editables para instructores
        if usuario.rol == "INSTRUCTOR":
            if "programa_formacion" in request.POST:
                usuario.programa_formacion = request.POST.get("programa_formacion", "").strip() or None
            if "ficha" in request.POST:
                usuario.ficha = request.POST.get("ficha", "").strip() or None

        # Foto de perfil
        if request.POST.get("eliminar_foto") == "1" and usuario.foto:
            import os

            if os.path.isfile(usuario.foto.path):
                os.remove(usuario.foto.path)
            usuario.foto = None
        elif "foto" in request.FILES:
            foto = request.FILES["foto"]
            if foto.size > 5 * 1024 * 1024:
                django_messages.error(request, "La foto no puede superar 5 MB.")
                return redirect("mi_perfil")
            if not foto.content_type.startswith("image/"):
                django_messages.error(request, "El archivo debe ser una imagen.")
                return redirect("mi_perfil")
            usuario.foto = foto

        usuario.save()
        django_messages.success(request, "Perfil actualizado correctamente.")
        return redirect("mi_perfil")

    # Contexto con información adicional según el rol
    from control_acceso.models import RegistroAcceso
    from django.utils import timezone

    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)

    total_accesos_mes = RegistroAcceso.objects.filter(
        usuario=usuario, tipo="INGRESO", fecha_hora_ingreso__date__gte=inicio_mes
    ).count()

    ultimo_acceso = (
        RegistroAcceso.objects.filter(usuario=usuario, tipo="INGRESO").order_by("-fecha_hora_ingreso").first()
    )

    # Estado actual para todos los roles
    registro_activo = (
        RegistroAcceso.objects.filter(
            usuario=usuario, tipo="INGRESO", fecha_hora_egreso__isnull=True, fecha_hora_ingreso__date=hoy
        )
        .order_by("-fecha_hora_ingreso")
        .first()
    )

    from usuarios.models import ProgramaFormacion as PFModel

    programas = list(PFModel.objects.filter(activo=True).values_list("nombre", flat=True))

    context = {
        "usuario": usuario,
        "total_accesos_mes": total_accesos_mes,
        "ultimo_acceso": ultimo_acceso,
        "en_centro": registro_activo is not None,
        "registro_activo": registro_activo,
        "programas": programas,
    }

    # Para instructores: cargar fichas disponibles por programa para el selector dinámico
    if usuario.rol == "INSTRUCTOR":
        from usuarios.models import Usuario as Usr
        import json

        fichas_qs = (
            Usr.objects.filter(rol="APRENDIZ", activo=True, ficha__isnull=False)
            .exclude(ficha="")
            .values("programa_formacion", "ficha")
            .distinct()
            .order_by("programa_formacion", "ficha")
        )
        fichas_por_programa = {}
        for item in fichas_qs:
            prog = item["programa_formacion"] or ""
            ficha = item["ficha"]
            fichas_por_programa.setdefault(prog, [])
            if ficha not in fichas_por_programa[prog]:
                fichas_por_programa[prog].append(ficha)
        context["fichas_por_programa_json"] = json.dumps(fichas_por_programa)

    return render(request, "perfil.html", context)


# ==============================================
# VISTAS ESPECÍFICAS PARA APRENDIZ
# ==============================================


@login_required
@rol_requerido("APRENDIZ")
def mi_asistencia_view(request):
    """
    Vista de asistencia del aprendiz
    """
    # Aquí puedes agregar lógica para obtener datos reales de asistencia
    from control_acceso.models import RegistroAcceso
    from django.utils import timezone

    usuario = request.user
    hoy = timezone.now().date()

    # Obtener registros de este mes
    inicio_mes = hoy.replace(day=1)
    registros_mes = RegistroAcceso.objects.filter(
        usuario=usuario, tipo="INGRESO", fecha_hora_ingreso__gte=inicio_mes
    ).order_by("-fecha_hora_ingreso")

    context = {
        "usuario": usuario,
        "registros_mes": registros_mes,
        "dias_asistidos": registros_mes.count(),
        "total_dias_mes": hoy.day,
    }

    return render(request, "dashboard/aprendiz/mi_asistencia.html", context)


@login_required
@excluir_visitantes
def mis_alertas_view(request):
    """
    Vista de alertas/notificaciones del usuario.
    Accesible por todos los roles excepto VISITANTE.
    Para BRIGADA incluye además el historial completo de incidentes del sistema.
    """
    from usuarios.models import Notificacion

    # Obtener notificaciones del usuario actual (todas, sin límite duro)
    notificaciones = Notificacion.objects.filter(destinatario=request.user)

    # Separar por estado de lectura (hasta 50 de cada tipo)
    no_leidas = notificaciones.filter(leida=False).order_by("-fecha_creacion")[:50]
    leidas = notificaciones.filter(leida=True).order_by("-fecha_creacion")[:50]

    # Contar totales reales
    total_no_leidas = notificaciones.filter(leida=False).count()
    total_leidas = notificaciones.filter(leida=True).count()
    total_notificaciones = notificaciones.count()

    context = {
        "no_leidas": no_leidas,
        "leidas": leidas,
        "total_no_leidas": total_no_leidas,
        "total_leidas": total_leidas,
        "total_notificaciones": total_notificaciones,
    }

    # Para BRIGADA: agregar historial completo de incidentes del sistema
    if request.user.rol == "BRIGADA":
        from reportes.models import Incidente
        from django.utils import timezone
        from datetime import timedelta

        todos_incidentes = Incidente.objects.select_related("reportado_por", "asignado_a").order_by("-fecha_reporte")[
            :100
        ]
        incidentes_pendientes = Incidente.objects.exclude(estado__in=["RESUELTO", "CERRADO"]).count()
        incidentes_criticos_sla = (
            Incidente.objects.exclude(estado__in=["RESUELTO", "CERRADO"])
            .filter(fecha_reporte__lte=timezone.now() - timedelta(hours=72))
            .count()
        )
        # Anotar SLA en cada incidente
        from reportes.views_incidentes import _calcular_sla

        for inc in todos_incidentes:
            inc.sla_estado = _calcular_sla(inc)
        context["todos_incidentes"] = todos_incidentes
        context["incidentes_pendientes_brigada"] = incidentes_pendientes
        context["incidentes_criticos_sla_brigada"] = incidentes_criticos_sla

    return render(request, "dashboard/aprendiz/mis_alertas.html", context)


# ==============================================
# FIN VISTAS PARA APRENDIZ
# ==============================================

# ==============================================
# VISTAS ESPECÍFICAS PARA INSTRUCTOR
# ==============================================


@login_required
@rol_requerido("INSTRUCTOR")
def mis_aprendices_view(request):
    """
    Vista para que el instructor vea los aprendices de SUS fichas asignadas.
    Si no tiene fichas asignadas, muestra un aviso para configurar el perfil.
    """
    from usuarios.models import Usuario
    from control_acceso.models import RegistroAcceso
    from django.db.models import OuterRef, Subquery, Exists
    from django.utils import timezone

    instructor = request.user
    ficha_instructor = instructor.ficha or ""  # única ficha asignada
    programa_instructor = instructor.programa_formacion or ""

    hoy = timezone.now().date()

    aprendices_con_asistencia = []
    total_aprendices = 0
    presentes_hoy = 0

    ficha_seleccionada = ficha_instructor  # solo puede ver su propia ficha

    if ficha_instructor:
        ultimo_registro_hoy = RegistroAcceso.objects.filter(
            usuario=OuterRef("pk"), fecha_hora_ingreso__date=hoy
        ).order_by("-fecha_hora_ingreso")

        aprendices = (
            Usuario.objects.filter(rol="APRENDIZ", activo=True, ficha=ficha_seleccionada)
            .annotate(
                tiene_registro_hoy=Exists(
                    RegistroAcceso.objects.filter(usuario=OuterRef("pk"), fecha_hora_ingreso__date=hoy)
                ),
                _en_centro=Exists(
                    RegistroAcceso.objects.filter(
                        usuario=OuterRef("pk"), fecha_hora_ingreso__date=hoy, fecha_hora_egreso__isnull=True
                    )
                ),
                _hora_ingreso=Subquery(ultimo_registro_hoy.values("fecha_hora_ingreso")[:1]),
            )
            .order_by("last_name", "first_name")
        )

        total_aprendices = aprendices.count()

        for aprendiz in aprendices:
            asistio = aprendiz.tiene_registro_hoy
            if asistio:
                presentes_hoy += 1
            aprendices_con_asistencia.append(
                {
                    "usuario": aprendiz,
                    "presente": aprendiz._en_centro,
                    "hora_ingreso": aprendiz._hora_ingreso,
                    "tiene_registro_hoy": asistio,
                }
            )

    ausentes = total_aprendices - presentes_hoy
    porcentaje_asistencia = round((presentes_hoy / total_aprendices) * 100) if total_aprendices > 0 else 0

    context = {
        "ficha_instructor": ficha_instructor,
        "programa_instructor": programa_instructor,
        "ficha_seleccionada": ficha_seleccionada,
        "aprendices": aprendices_con_asistencia,
        "total_aprendices": total_aprendices,
        "presentes_hoy": presentes_hoy,
        "ausentes": ausentes,
        "porcentaje_asistencia": porcentaje_asistencia,
    }
    return render(request, "dashboard/instructor/mis_aprendices.html", context)


@login_required
@rol_requerido("INSTRUCTOR")
def evacuacion_instructor_view(request):
    """
    Vista para que el instructor tome lista de evacuación de sus aprendices
    cuando hay una emergencia de alerta masiva activa.
    """
    from usuarios.models import Usuario
    from emergencias.models import Emergencia, RegistroEvacuacion

    instructor = request.user
    fichas = instructor.get_fichas_list() or ([instructor.ficha] if instructor.ficha else [])

    # Buscar emergencia de alerta masiva activa
    emergencia = (
        Emergencia.objects.filter(
            tipo__alerta_masiva=True,
            estado__in=["REPORTADA", "EN_ATENCION"],
        )
        .order_by("-fecha_hora_reporte")
        .first()
    )

    # Si no hay emergencia masiva activa, redirigir al inicio
    if not emergencia:
        from django.contrib import messages

        messages.info(request, "No hay ninguna emergencia masiva activa en este momento.")
        return redirect("dashboard")

    aprendices_data = []
    if fichas:
        from control_acceso.models import RegistroAcceso
        from django.db.models import Exists, OuterRef
        from django.utils import timezone

        hoy = timezone.now().date()

        # Solo aprendices que registraron asistencia hoy (están en el centro)
        aprendices = (
            Usuario.objects.filter(rol="APRENDIZ", activo=True, ficha__in=fichas)
            .filter(
                Exists(
                    RegistroAcceso.objects.filter(
                        usuario=OuterRef("pk"),
                        fecha_hora_ingreso__date=hoy,
                    )
                )
            )
            .order_by("ficha", "last_name", "first_name")
        )

        registros = {
            r.usuario_id: r for r in RegistroEvacuacion.objects.filter(emergencia=emergencia, usuario__in=aprendices)
        }

        for aprendiz in aprendices:
            reg = registros.get(aprendiz.id)
            aprendices_data.append(
                {
                    "usuario": aprendiz,
                    "confirmado": reg.confirmado if reg else False,
                    "ausente": reg.ausente if reg else False,
                }
            )

    context = {
        "emergencia": emergencia,
        "aprendices": aprendices_data,
        "fichas": fichas,
        "total": len(aprendices_data),
        "confirmados": sum(1 for a in aprendices_data if a["confirmado"]),
        "ausentes": sum(1 for a in aprendices_data if a["ausente"]),
    }
    return render(request, "dashboard/instructor/evacuacion.html", context)


# ==============================================
# VISTAS ESPECÍFICAS PARA ADMINISTRATIVO
# ==============================================


@login_required
@rol_requerido("ADMINISTRATIVO", "COORDINADOR_SST")
def gestion_usuarios_view(request):
    """
    Vista para gestión de usuarios (Administrativo) con paginación
    """
    from usuarios.models import Usuario
    from django.core.paginator import Paginator

    from usuarios.models import ProgramaFormacion as PFModel

    # Excluir superusuarios para seguridad básica en la vista
    usuarios = Usuario.objects.all().exclude(is_superuser=True).order_by("-fecha_registro")

    # Paginación: 20 usuarios por página
    paginator = Paginator(usuarios, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    programas = list(PFModel.objects.filter(activo=True).values_list("nombre", flat=True))
    context = {
        "usuarios": page_obj,
        "page_obj": page_obj,
        "total_usuarios": usuarios.count(),
        "programas": programas,
    }
    return render(request, "dashboard/administrativo/gestion_usuarios.html", context)


@login_required
@rol_requerido("ADMINISTRATIVO", "COORDINADOR_SST")
def configuracion_view(request):
    """
    Vista de configuración del sistema
    """
    from control_acceso.models import ConfiguracionAforo
    from django.contrib import messages as django_messages

    config_aforo = ConfiguracionAforo.objects.filter(activo=True).first()

    if request.method == "POST":
        try:
            aforo_maximo = int(request.POST.get("aforo_maximo", 2000))
            umbral_alerta = int(request.POST.get("umbral_alerta", 90))
            mensaje_alerta = request.POST.get("mensaje_alerta", "").strip()
            aforo_activo = request.POST.get("aforo_activo") == "on"
            aforo_minimo = round(aforo_maximo * umbral_alerta / 100)

            if config_aforo:
                config_aforo.aforo_maximo = aforo_maximo
                config_aforo.aforo_minimo = aforo_minimo
                config_aforo.mensaje_alerta = mensaje_alerta
                config_aforo.activo = aforo_activo
                config_aforo.save()
            else:
                ConfiguracionAforo.objects.create(
                    aforo_maximo=aforo_maximo,
                    aforo_minimo=aforo_minimo,
                    mensaje_alerta=mensaje_alerta,
                    activo=aforo_activo,
                )

            django_messages.success(request, "Configuración guardada correctamente.")
        except (ValueError, TypeError):
            django_messages.error(request, "Error al guardar: verifique los valores ingresados.")
        return redirect("configuracion_sistema")

    # Calcular umbral como porcentaje para el slider
    umbral_alerta = 90
    if config_aforo and config_aforo.aforo_maximo:
        umbral_alerta = round(config_aforo.aforo_minimo / config_aforo.aforo_maximo * 100)

    context = {
        "config_aforo": config_aforo,
        "umbral_alerta": umbral_alerta,
    }
    return render(request, "dashboard/administrativo/configuracion.html", context)


# ==============================================
# VISTAS ESPECÍFICAS PARA VIGILANCIA
# ==============================================


@login_required
@rol_requerido("VIGILANCIA", "ADMINISTRATIVO")
def gestion_visitantes_view(request):
    """
    Vista para gestión de visitantes
    PERMISOS: VIGILANCIA y ADMINISTRATIVO
    """
    from usuarios.models import Visitante, Usuario
    from django.utils import timezone
    from django.contrib import messages as django_messages

    hoy = timezone.now().date()

    if request.method == "POST" and request.POST.get("accion") == "salida":
        from django.utils import timezone as tz

        visitante_id = request.POST.get("visitante_id")
        try:
            v = Visitante.objects.get(id=visitante_id)
            v.hora_salida = tz.localtime(tz.now()).time()
            v.save(update_fields=["hora_salida"])
            django_messages.success(request, f"Salida de {v.nombre_completo} registrada.")
        except Visitante.DoesNotExist:
            django_messages.error(request, "Visitante no encontrado.")
        return redirect("gestion_visitantes")

    if request.method == "POST":
        nombre_completo = request.POST.get("nombre_completo", "").strip()
        tipo_documento = request.POST.get("tipo_documento", "CC")
        numero_documento = request.POST.get("numero_documento", "").strip()
        email = request.POST.get("email", "").strip()
        telefono = request.POST.get("telefono", "").strip()
        entidad = request.POST.get("entidad", "").strip()
        motivo_visita = request.POST.get("motivo_visita", "").strip()

        if not nombre_completo or not numero_documento or not email or not motivo_visita:
            django_messages.error(request, "Complete todos los campos obligatorios.")
        else:
            visitante = Visitante.objects.create(
                nombre_completo=nombre_completo,
                tipo_documento=tipo_documento,
                numero_documento=numero_documento,
                email=email,
                telefono=telefono,
                entidad=entidad,
                motivo_visita=motivo_visita,
                registrado_por=request.user,
            )

            # Crear o reactivar cuenta de usuario visitante
            import logging
            import threading

            from django.core.mail import send_mail

            logger = logging.getLogger(__name__)

            partes = nombre_completo.split()
            first_name = partes[0]
            last_name = " ".join(partes[1:]) if len(partes) > 1 else ""

            # Buscar cuenta existente por email (más confiable que documento)
            cuenta = Usuario.objects.filter(email__iexact=email, rol="VISITANTE").first()

            if not cuenta:
                # Buscar por documento con rol VISITANTE
                cuenta = Usuario.objects.filter(numero_documento=numero_documento, rol="VISITANTE").first()

            if cuenta:
                # Reactivar cuenta existente
                cuenta.email = email
                cuenta.activo = True
                cuenta.is_active = True
                cuenta.estado_cuenta = "ACTIVO"
                cuenta.set_unusable_password()
                cuenta.save(update_fields=["email", "activo", "is_active", "estado_cuenta", "password"])
            else:
                # Generar username único y número de documento sin conflictos
                import hashlib

                doc_visitante = f"v{numero_documento}"
                # Si ese documento ya existe en Usuario (para otro rol), usar hash del email
                if (
                    Usuario.objects.filter(numero_documento=doc_visitante).exists()
                    or Usuario.objects.filter(numero_documento=numero_documento).exists()
                ):
                    doc_visitante = "v" + hashlib.md5(email.encode()).hexdigest()[:12]

                username = f"visitante_{numero_documento}"
                if Usuario.objects.filter(username=username).exists():
                    username = f"visitante_{hashlib.md5(email.encode()).hexdigest()[:10]}"

                cuenta = Usuario.objects.create_user(
                    username=username,
                    email=email,
                    password=None,
                    first_name=first_name,
                    last_name=last_name,
                    rol="VISITANTE",
                    tipo_documento=tipo_documento,
                    numero_documento=doc_visitante,
                    telefono=telefono,
                    activo=True,
                    is_active=True,
                    estado_cuenta="ACTIVO",
                )
                cuenta.set_unusable_password()
                cuenta.save(update_fields=["password"])

            visitante.usuario = cuenta
            visitante.save(update_fields=["usuario"])

            def _enviar_correo():
                try:
                    send_mail(
                        subject="Tu acceso al Sistema SST - Centro Minero SENA",
                        message=(
                            f"Hola {nombre_completo},\n\n"
                            f"Has sido registrado como visitante en el Centro Minero SENA.\n\n"
                            f"Puedes ingresar al sistema usando el siguiente enlace:\n"
                            f"http://127.0.0.1:8000/accounts/login/visitante/\n\n"
                            f"Solo ingresa tu correo electrónico ({email}) y podrás acceder.\n\n"
                            f"Ten en cuenta que este acceso estará disponible únicamente durante el día de hoy.\n\n"
                            f"Saludos,\nSistema SST - Centro Minero SENA"
                        ),
                        from_email=None,
                        recipient_list=[email],
                        fail_silently=False,
                    )
                    logger.info(f"Correo enviado a visitante: {email}")
                except Exception as e:
                    logger.error(f"Error enviando correo a visitante {email}: {e}")

            threading.Thread(target=_enviar_correo, daemon=True).start()
            django_messages.success(
                request, f"Visitante {nombre_completo} registrado. Se envió el enlace de acceso a {email}."
            )

        return redirect("gestion_visitantes")

    visitantes_hoy = Visitante.objects.filter(fecha_visita=hoy).order_by("-hora_ingreso")
    context = {
        "visitantes": visitantes_hoy,
        "total_hoy": visitantes_hoy.count(),
        "activos_ahora": visitantes_hoy.filter(hora_salida__isnull=True).count(),
    }
    return render(request, "dashboard/vigilancia/gestion_visitantes.html", context)


@login_required
@rol_requerido("BRIGADA")
def equipos_brigada_view(request):
    """
    Vista para gestión de equipos de emergencia (Brigada + Coordinador SST).
    GET  → lista con filtros
    POST → CRUD (crear, editar, eliminar) — solo COORDINADOR_SST
    """
    from mapas.models import EquipamientoSeguridad, EdificioBloque
    from django.db.models import Q
    from django.utils import timezone
    from datetime import timedelta
    from django.contrib import messages as django_messages

    # ── CRUD (solo coordinador) ──────────────────────────────────────
    if request.method == "POST":
        if request.user.rol != "BRIGADA":
            django_messages.error(request, "No tienes permiso para realizar esta acción.")
            return redirect("equipos_brigada")

        accion = request.POST.get("accion", "")

        if accion == "crear":
            tipo = request.POST.get("tipo", "").strip()
            codigo = request.POST.get("codigo", "").strip()
            descripcion = request.POST.get("descripcion", "").strip()
            estado = request.POST.get("estado", "OPERATIVO")
            edificio_id = request.POST.get("edificio_id") or None
            piso = request.POST.get("piso", 1)
            proxima_revision = request.POST.get("proxima_revision") or None

            if not tipo or not codigo:
                django_messages.error(request, "El tipo y el código son obligatorios.")
            elif EquipamientoSeguridad.objects.filter(codigo__iexact=codigo).exists():
                django_messages.error(request, f'Ya existe un equipo con el código "{codigo}".')
            else:
                edificio = None
                if edificio_id:
                    try:
                        edificio = EdificioBloque.objects.get(pk=edificio_id)
                    except EdificioBloque.DoesNotExist:
                        pass
                lat = edificio.latitud if edificio else 5.8520
                lng = edificio.longitud if edificio else -73.0318
                EquipamientoSeguridad.objects.create(
                    tipo=tipo,
                    codigo=codigo,
                    nombre=descripcion or codigo,
                    descripcion=descripcion,
                    estado=estado,
                    edificio=edificio,
                    piso=int(piso),
                    latitud=lat,
                    longitud=lng,
                    fecha_instalacion=timezone.now(),
                    proxima_revision=proxima_revision or None,
                )
                django_messages.success(request, f'Equipo "{codigo}" registrado correctamente.')

        elif accion == "editar":
            equipo_id = request.POST.get("equipo_id")
            try:
                equipo = EquipamientoSeguridad.objects.get(pk=equipo_id)
                codigo_nuevo = request.POST.get("codigo", "").strip()
                if not codigo_nuevo:
                    django_messages.error(request, "El código no puede estar vacío.")
                elif EquipamientoSeguridad.objects.filter(codigo__iexact=codigo_nuevo).exclude(pk=equipo_id).exists():
                    django_messages.error(request, f'Ya existe otro equipo con el código "{codigo_nuevo}".')
                else:
                    edificio_id = request.POST.get("edificio_id") or None
                    edificio = None
                    if edificio_id:
                        try:
                            edificio = EdificioBloque.objects.get(pk=edificio_id)
                        except EdificioBloque.DoesNotExist:
                            pass
                    descripcion = request.POST.get("descripcion", "").strip()
                    equipo.tipo = request.POST.get("tipo", equipo.tipo)
                    equipo.codigo = codigo_nuevo
                    equipo.nombre = descripcion or codigo_nuevo
                    equipo.descripcion = descripcion
                    equipo.estado = request.POST.get("estado", equipo.estado)
                    equipo.edificio = edificio
                    equipo.piso = int(request.POST.get("piso", equipo.piso))
                    equipo.proxima_revision = request.POST.get("proxima_revision") or None
                    if edificio:
                        equipo.latitud = edificio.latitud
                        equipo.longitud = edificio.longitud
                    equipo.save()
                    django_messages.success(request, f'Equipo "{codigo_nuevo}" actualizado.')
            except EquipamientoSeguridad.DoesNotExist:
                django_messages.error(request, "Equipo no encontrado.")

        elif accion == "eliminar":
            equipo_id = request.POST.get("equipo_id")
            try:
                equipo = EquipamientoSeguridad.objects.get(pk=equipo_id)
                codigo = equipo.codigo
                equipo.delete()
                django_messages.success(request, f'Equipo "{codigo}" eliminado.')
            except EquipamientoSeguridad.DoesNotExist:
                django_messages.error(request, "Equipo no encontrado.")

        return redirect("equipos_brigada")

    # ── GET ──────────────────────────────────────────────────────────
    tipo_filtro = request.GET.get("tipo", "")
    estado_filtro = request.GET.get("estado", "")
    busqueda = request.GET.get("q", "")

    equipos = EquipamientoSeguridad.objects.select_related("edificio", "responsable").all()

    if tipo_filtro:
        equipos = equipos.filter(tipo=tipo_filtro)
    if estado_filtro:
        equipos = equipos.filter(estado=estado_filtro)
    if busqueda:
        equipos = equipos.filter(
            Q(nombre__icontains=busqueda) | Q(codigo__icontains=busqueda) | Q(edificio__nombre__icontains=busqueda)
        )

    equipos = equipos.order_by("-fecha_creacion")

    total_equipos = EquipamientoSeguridad.objects.count()
    operativos = EquipamientoSeguridad.objects.filter(estado="OPERATIVO").count()
    en_mantenimiento = EquipamientoSeguridad.objects.filter(estado="MANTENIMIENTO").count()
    fuera_servicio = EquipamientoSeguridad.objects.filter(estado="FUERA_SERVICIO").count()

    fecha_limite = timezone.now() + timedelta(days=30)
    proximos_revision = EquipamientoSeguridad.objects.filter(
        proxima_revision__lte=fecha_limite, proxima_revision__gte=timezone.now()
    ).count()

    tipos_equipamiento = EquipamientoSeguridad.TIPO_EQUIPAMIENTO
    edificios = EdificioBloque.objects.filter(activo=True).order_by("nombre")

    context = {
        "equipos": equipos,
        "total_equipos": total_equipos,
        "operativos": operativos,
        "en_mantenimiento": en_mantenimiento,
        "fuera_servicio": fuera_servicio,
        "proximos_revision": proximos_revision,
        "tipos_equipamiento": tipos_equipamiento,
        "tipo_filtro": tipo_filtro,
        "estado_filtro": estado_filtro,
        "busqueda": busqueda,
        "edificios": edificios,
        "now": timezone.now(),
        "estados": [
            ("OPERATIVO", "Operativo"),
            ("MANTENIMIENTO", "En mantenimiento"),
            ("FUERA_SERVICIO", "Fuera de servicio"),
        ],
    }
    return render(request, "dashboard/brigada/equipos.html", context)


@login_required
@rol_requerido("COORDINADOR_SST")
def programas_formacion_view(request):
    """
    CRUD de programas de formación — solo Coordinador SST.
    GET  → lista todos los programas
    POST → crea, edita o elimina según el campo 'accion'
    """
    from usuarios.models import ProgramaFormacion
    from django.contrib import messages as django_messages

    if request.method == "POST":
        accion = request.POST.get("accion", "")

        if accion == "crear":
            nombre = request.POST.get("nombre", "").strip()
            if not nombre:
                django_messages.error(request, "El nombre del programa no puede estar vacío.")
            elif ProgramaFormacion.objects.filter(nombre__iexact=nombre).exists():
                django_messages.error(request, f'Ya existe un programa llamado "{nombre}".')
            else:
                ProgramaFormacion.objects.create(nombre=nombre)
                django_messages.success(request, f'Programa "{nombre}" creado correctamente.')

        elif accion == "editar":
            programa_id = request.POST.get("programa_id")
            nombre = request.POST.get("nombre", "").strip()
            try:
                prog = ProgramaFormacion.objects.get(pk=programa_id)
                if not nombre:
                    django_messages.error(request, "El nombre no puede estar vacío.")
                elif ProgramaFormacion.objects.filter(nombre__iexact=nombre).exclude(pk=programa_id).exists():
                    django_messages.error(request, f'Ya existe un programa llamado "{nombre}".')
                else:
                    prog.nombre = nombre
                    prog.save()
                    django_messages.success(request, f'Programa actualizado a "{nombre}".')
            except ProgramaFormacion.DoesNotExist:
                django_messages.error(request, "Programa no encontrado.")

        elif accion == "eliminar":
            programa_id = request.POST.get("programa_id")
            try:
                prog = ProgramaFormacion.objects.get(pk=programa_id)
                # Verificar que no haya usuarios usando este programa
                from usuarios.models import Usuario

                en_uso = Usuario.objects.filter(programa_formacion__iexact=prog.nombre).exists()
                if en_uso:
                    django_messages.error(
                        request,
                        f'No se puede eliminar "{prog.nombre}" porque hay usuarios asignados a este programa.',
                    )
                else:
                    nombre = prog.nombre
                    prog.delete()
                    django_messages.success(request, f'Programa "{nombre}" eliminado.')
            except ProgramaFormacion.DoesNotExist:
                django_messages.error(request, "Programa no encontrado.")

        return redirect("programas_formacion")

    programas = ProgramaFormacion.objects.all().order_by("-fecha_creacion")
    return render(request, "dashboard/coordinador/programas_formacion.html", {"programas": programas})


@login_required
@rol_requerido("COORDINADOR_SST")
def fichas_formacion_view(request):
    """
    CRUD de fichas por programa de formación — solo Coordinador SST.
    """
    from usuarios.models import FichaFormacion, ProgramaFormacion
    from django.contrib import messages as django_messages

    if request.method == "POST":
        accion = request.POST.get("accion")

        if accion == "crear":
            programa_id = request.POST.get("programa_id")
            numero = request.POST.get("numero", "").strip()
            if not numero:
                django_messages.error(request, "El número de ficha no puede estar vacío.")
            elif not numero.isdigit() or len(numero) != 7:
                django_messages.error(request, "El número de ficha debe tener exactamente 7 dígitos numéricos.")
            elif not programa_id:
                django_messages.error(request, "Debes seleccionar un programa.")
            else:
                try:
                    programa = ProgramaFormacion.objects.get(pk=programa_id)
                    if FichaFormacion.objects.filter(programa=programa, numero=numero).exists():
                        django_messages.error(request, f'La ficha "{numero}" ya existe en ese programa.')
                    else:
                        FichaFormacion.objects.create(programa=programa, numero=numero)
                        django_messages.success(request, f'Ficha "{numero}" creada en "{programa.nombre}".')
                except ProgramaFormacion.DoesNotExist:
                    django_messages.error(request, "Programa no encontrado.")

        elif accion == "eliminar":
            ficha_id = request.POST.get("ficha_id")
            try:
                ficha = FichaFormacion.objects.get(pk=ficha_id)
                numero = ficha.numero
                ficha.delete()
                django_messages.success(request, f'Ficha "{numero}" eliminada.')
            except FichaFormacion.DoesNotExist:
                django_messages.error(request, "Ficha no encontrada.")

        return redirect("fichas_formacion")

    from django.db.models import Prefetch

    programas = (
        ProgramaFormacion.objects.filter(activo=True)
        .prefetch_related(Prefetch("fichas", queryset=FichaFormacion.objects.order_by("-fecha_creacion")))
        .order_by("-fecha_creacion")
    )
    return render(request, "dashboard/coordinador/fichas_formacion.html", {"programas": programas})


def api_fichas_por_programa(request):
    """
    JSON público: fichas activas para un programa dado.
    GET /api/fichas/?programa_id=X  o  ?programa_nombre=Nombre
    Usado en el formulario de registro (sin autenticación).
    """
    from usuarios.models import FichaFormacion, ProgramaFormacion
    from django.http import JsonResponse

    programa_id = request.GET.get("programa_id")
    programa_nombre = request.GET.get("programa_nombre", "").strip()

    try:
        if programa_id:
            programa = ProgramaFormacion.objects.get(pk=programa_id, activo=True)
        elif programa_nombre:
            programa = ProgramaFormacion.objects.get(nombre__iexact=programa_nombre, activo=True)
        else:
            return JsonResponse({"fichas": []})
    except ProgramaFormacion.DoesNotExist:
        return JsonResponse({"fichas": []})

    fichas = list(
        FichaFormacion.objects.filter(programa=programa, activo=True)
        .order_by("numero")
        .values_list("numero", flat=True)
    )
    return JsonResponse({"fichas": fichas})


@login_required
@rol_requerido("BRIGADA")
def mi_brigada_view(request):
    """
    Vista para ver los miembros de la brigada
    """
    from usuarios.models import Usuario

    miembros = Usuario.objects.filter(rol="BRIGADA", activo=True).exclude(id=request.user.id)
    context = {
        "miembros": miembros,
        "total_miembros": miembros.count() + 1,  # Contando al usuario actual
    }
    return render(request, "dashboard/brigada/mi_brigada.html", context)


# ==============================================


# Vistas para usuarios autenticados (usan base.html)
@rol_requerido("ADMINISTRATIVO", "VIGILANCIA")
def control_acceso_view(request):
    """
    Vista de Control de Acceso
    PERMISOS: Solo ADMINISTRATIVO y VIGILANCIA
    """
    return render(request, "control_acceso.html")


# Importar las vistas de mapas
from mapas.views import mapa_interactivo, plano_centro as plano_centro_view


@rol_requerido("ADMINISTRATIVO", "BRIGADA")
def emergencias_view(request):
    """
    Vista de Emergencias
    PERMISOS: Solo ADMINISTRATIVO, BRIGADA y VIGILANCIA
    """
    return render(request, "emergencias.html")


@excluir_visitantes
def reportes_view(request):
    """
    Vista de Reportes
    PERMISOS: Todos excepto VISITANTE
    """
    return render(request, "reportes/index.html")


# ==============================================
# EXPORTAR ACCESOS A EXCEL
# ==============================================


@login_required
@rol_requerido("ADMINISTRATIVO", "VIGILANCIA", "INSTRUCTOR")
def exportar_accesos_excel(request):
    """
    Exporta los registros de acceso del día (o rango indicado) a Excel.
    Parámetros GET: fecha_desde, fecha_hasta, rol
    """
    from control_acceso.models import RegistroAcceso
    from django.utils import timezone
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    hoy = timezone.now().date()

    fecha_desde_str = request.GET.get("fecha_desde", hoy.isoformat())
    fecha_hasta_str = request.GET.get("fecha_hasta", hoy.isoformat())
    rol_filtro = request.GET.get("rol", "")

    try:
        fecha_desde = datetime.strptime(fecha_desde_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        fecha_desde = hoy
    try:
        fecha_hasta = datetime.strptime(fecha_hasta_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        fecha_hasta = hoy

    registros = (
        RegistroAcceso.objects.select_related("usuario")
        .filter(
            tipo="INGRESO",
            fecha_hora_ingreso__date__gte=fecha_desde,
            fecha_hora_ingreso__date__lte=fecha_hasta,
        )
        .order_by("-fecha_hora_ingreso")
    )

    if rol_filtro:
        registros = registros.filter(usuario__rol=rol_filtro)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros de Acceso"

    verde = PatternFill("solid", fgColor="39A900")
    verde_oscuro = PatternFill("solid", fgColor="2E7D32")
    borde = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # Título
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"Registros de Acceso — Centro Minero SENA — {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = verde
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Encabezados
    headers = [
        "Nombre Completo",
        "Documento",
        "Rol",
        "Ficha / Programa",
        "Hora Ingreso",
        "Hora Egreso",
        "Método",
        "Estado",
    ]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = verde_oscuro
        cell.alignment = Alignment(horizontal="center")
        cell.border = borde
    ws.row_dimensions[2].height = 15

    # Datos
    for i, reg in enumerate(registros, 3):
        if reg.fecha_hora_egreso:
            estado = "Salió"
        else:
            estado = "En centro"

        row = [
            reg.usuario.get_full_name() or reg.usuario.username,
            reg.usuario.numero_documento or "",
            reg.usuario.get_rol_display(),
            f"{reg.usuario.ficha or ''} {reg.usuario.programa_formacion or ''}".strip() or "—",
            reg.fecha_hora_ingreso.strftime("%d/%m/%Y %H:%M"),
            reg.fecha_hora_egreso.strftime("%d/%m/%Y %H:%M") if reg.fecha_hora_egreso else "—",
            reg.get_metodo_ingreso_display(),
            estado,
        ]
        ws.append(row)
        fill = PatternFill("solid", fgColor="F9FBF9") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for cell in ws[i]:
            cell.fill = fill
            cell.border = borde
            cell.alignment = Alignment(vertical="center")

    # Anchos de columna
    for col, ancho in zip("ABCDEFGH", [28, 15, 16, 30, 18, 18, 12, 12]):
        ws.column_dimensions[col].width = ancho

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    fecha_str = f"{fecha_desde.strftime('%Y%m%d')}_{fecha_hasta.strftime('%Y%m%d')}"
    response["Content-Disposition"] = f'attachment; filename="accesos_{fecha_str}.xlsx"'
    wb.save(response)
    return response


# ==============================================
# HISTORIAL DE ACCESOS (vista HTML)
# ==============================================


@login_required
@rol_requerido("ADMINISTRATIVO", "VIGILANCIA", "INSTRUCTOR")
def historial_accesos_view(request):
    """
    Vista HTML con historial de accesos filtrable + exportación Excel.
    """
    from control_acceso.models import RegistroAcceso
    from django.core.paginator import Paginator
    from django.utils import timezone
    from datetime import datetime

    hoy = timezone.now().date()
    fecha_desde_str = request.GET.get("fecha_desde", hoy.isoformat())
    fecha_hasta_str = request.GET.get("fecha_hasta", hoy.isoformat())
    rol_filtro = request.GET.get("rol", "")
    q = request.GET.get("q", "").strip()

    try:
        fecha_desde = datetime.strptime(fecha_desde_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        fecha_desde = hoy
    try:
        fecha_hasta = datetime.strptime(fecha_hasta_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        fecha_hasta = hoy

    registros = (
        RegistroAcceso.objects.select_related("usuario")
        .filter(
            tipo="INGRESO",
            fecha_hora_ingreso__date__gte=fecha_desde,
            fecha_hora_ingreso__date__lte=fecha_hasta,
        )
        .order_by("-fecha_hora_ingreso")
    )

    if rol_filtro:
        registros = registros.filter(usuario__rol=rol_filtro)
    if q:
        from django.db.models import Q

        registros = registros.filter(
            Q(usuario__first_name__icontains=q)
            | Q(usuario__last_name__icontains=q)
            | Q(usuario__numero_documento__icontains=q)
        )

    total = registros.count()
    paginator = Paginator(registros, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    from usuarios.models import Usuario as Usr

    roles_disponibles = Usr.objects.values_list("rol", flat=True).distinct().order_by("rol")

    context = {
        "registros": page_obj,
        "page_obj": page_obj,
        "total": total,
        "fecha_desde": fecha_desde_str,
        "fecha_hasta": fecha_hasta_str,
        "rol_filtro": rol_filtro,
        "q": q,
        "roles_disponibles": roles_disponibles,
        "hoy": hoy.isoformat(),
    }
    return render(request, "dashboard/vigilancia/historial_accesos.html", context)


handler404 = "sst_proyecto.error_views.error_404"
handler500 = "sst_proyecto.error_views.error_500"

urlpatterns = [
    # Chrome DevTools auto-discovery (silenciar 404 en logs)
    path(
        ".well-known/appspecific/com.chrome.devtools.json",
        lambda r: HttpResponse("{}", content_type="application/json"),
    ),
    # PWA: Service Worker y Manifest servidos desde la raíz
    path("sw.js", sw_view, name="sw"),
    path("manifest.json", manifest_view, name="manifest"),
    path("admin/", admin.site.urls),
    # Autenticación - Rutas principales (con debugging temporal)
    path("accounts/login/", custom_login_view, name="login"),
    path("accounts/login/visitante/", visitante_login_view, name="visitante_login"),
    path("accounts/logout/", auth_views.LogoutView.as_view(next_page="login"), name="logout"),
    # Recuperar Clave
    path(
        "accounts/password-reset/",
        auth_views.PasswordResetView.as_view(
            template_name="registration/recuperar_clave.html",
            email_template_name="registration/email_recuperacion.txt",
            html_email_template_name="registration/email_recuperacion.html",
            subject_template_name="registration/asunto_email.txt",
        ),
        name="password_reset",
    ),
    path(
        "accounts/password-reset/done/",
        auth_views.PasswordResetDoneView.as_view(template_name="registration/correo_enviado.html"),
        name="password_reset_done",
    ),
    path(
        "accounts/reset/<uidb64>/<token>/",
        auth_views.PasswordResetConfirmView.as_view(template_name="registration/nueva_clave.html"),
        name="password_reset_confirm",
    ),
    path(
        "accounts/reset/done/",
        auth_views.PasswordResetCompleteView.as_view(template_name="registration/clave_cambiada.html"),
        name="password_reset_complete",
    ),
    # Vistas principales del sistema (HTML templates)
    path("", dashboard_view, name="dashboard"),
    path("dashboard/", RedirectView.as_view(url="/", permanent=False)),
    path("perfil/", mi_perfil_view, name="mi_perfil"),
    path("acceso/", control_acceso_view, name="control_acceso"),
    path("mapas/", mapa_interactivo, name="mapas"),
    path("mapas/plano/", plano_centro_view, name="plano_centro"),
    path("emergencias/", emergencias_view, name="emergencias"),
    # ==============================================
    # URLs ESPECÍFICAS PARA APRENDIZ
    # ==============================================
    path("aprendiz/mis-accesos/", mi_asistencia_view, name="mi_asistencia"),
    path("alertas/", mis_alertas_view, name="mis_alertas"),
    # ==============================================
    # URLs PARA INSTRUCTOR
    path("instructor/mis-aprendices/", mis_aprendices_view, name="mis_aprendices"),
    path("instructor/evacuacion/", evacuacion_instructor_view, name="evacuacion_instructor"),
    # URLs PARA ADMINISTRATIVO
    path("administrativo/usuarios/", gestion_usuarios_view, name="gestion_usuarios"),
    path("administrativo/configuracion/", configuracion_view, name="configuracion_sistema"),
    # URLs PARA COORDINADOR SST
    path("coordinador/programas-formacion/", programas_formacion_view, name="programas_formacion"),
    path("coordinador/fichas-formacion/", fichas_formacion_view, name="fichas_formacion"),
    path("api/fichas/", api_fichas_por_programa, name="api_fichas_por_programa"),
    # URLs PARA VIGILANCIA
    path("vigilancia/visitantes/", gestion_visitantes_view, name="gestion_visitantes"),
    path("acceso/historial/", historial_accesos_view, name="historial_accesos"),
    path("acceso/exportar/excel/", exportar_accesos_excel, name="exportar_accesos_excel"),
    # URLs PARA BRIGADA
    path("brigada/equipos/", equipos_brigada_view, name="equipos_brigada"),
    path("brigada/mi-brigada/", mi_brigada_view, name="mi_brigada"),
    # Vista General de Reportes
    path("reportes/general/", reportes_view, name="reportes_general"),
    # ── APIs REST v1 (rutas canónicas versionadas) ─────────────────
    path("api/v1/auth/", include("usuarios.urls")),
    path("api/v1/acceso/", include("control_acceso.urls")),
    path("api/v1/mapas/", include("mapas.urls")),
    path("api/v1/emergencias/", include("emergencias.urls")),
    # ── APIs REST legacy (alias sin versión — mantiene compatibilidad) ─
    path("api/auth/", include("usuarios.urls")),
    path("api/acceso/", include("control_acceso.urls")),
    path("api/mapas/", include("mapas.urls")),
    path("api/emergencias/", include("emergencias.urls")),
    # Módulo de reportes (incluye vistas HTML de incidentes y API)
    path("reportes/", include("reportes.urls")),
    # ── Documentación automática de API (solo staff/admin) ──────────
    path("api/schema/", SpectacularAPIView.as_view(), name="api-schema"),
    path("api/schema/swagger/", SpectacularSwaggerView.as_view(url_name="api-schema"), name="swagger-ui"),
    path("api/schema/redoc/", SpectacularRedocView.as_view(url_name="api-schema"), name="redoc"),
]

if settings.DEBUG:
    urlpatterns += static(settings.MEDIA_URL, document_root=settings.MEDIA_ROOT)
    urlpatterns += static(settings.STATIC_URL, document_root=settings.STATIC_ROOT)
